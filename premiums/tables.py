import openpyxl
import MySQLdb

# 
"""
loading the company names
"""
# workbook = openpyxl.load_workbook('company.xlsx')
# sheet = workbook.active


# db = MySQLdb.Connect(host='localhost', user='root', passwd='admin', db='LULUPOP')
# cursor = db.cursor()

# for row in sheet.iter_rows(min_row=2, values_only=True):
#     company_name = row[0]
#     sql = "INSERT INTO companies (company_name) VALUES (%s)"
#     cursor.execute(sql, (company_name,))

# db.commit()
# db.close()


"""
loading the claim types
"""
workbook = openpyxl.load_workbook('claimtype.xlsx')
sheet = workbook.active


db = MySQLdb.Connect(host='localhost', user='root', passwd='admin', db='LULUPOP')
cursor = db.cursor()

for row in sheet.iter_rows(min_row=2, values_only=True):
     claim_type = row[0]
     sql = "INSERT INTO claim_type (claim_type) VALUES (%s)"
     cursor.execute(sql, (claim_type,))

db.commit()
db.close()


